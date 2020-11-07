#! /usr/bin/env python3

# daily_availability_reports -- tell deployment team who is available (or not)

import argparse
import logging
import os
import os.path
import re
import datetime
import time
import pprint
import json
import io
import csv
import sys
import random

import requests
import requests_html
import dotenv
import xlrd
import openpyxl

from http.cookiejar import LWPCookieJar, Cookie

import http.client
#http.client.HTTPConnection.debuglevel = 1


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import O365

import init_logging
from vc_session import get_session
log = logging.getLogger(__name__)
import config_avail as config_static


def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    config = init_config()

    # initialize office 365 graph api
    credentials = (config.CLIENT_ID, config.CLIENT_SECRET)

    scopes = [
            'https://graph.microsoft.com/Files.ReadWrite.All',
            'https://graph.microsoft.com/Mail.Read',
            'https://graph.microsoft.com/Mail.Read.Shared',
            'https://graph.microsoft.com/Mail.Send',
            'https://graph.microsoft.com/Mail.Send.Shared',
            'https://graph.microsoft.com/offline_access',
            'https://graph.microsoft.com/User.Read',
            'https://graph.microsoft.com/User.ReadBasic.All',
            #'https://microsoft.sharepoint-df.com/AllSites.Read',
            #'https://microsoft.sharepoint-df.com/MyFiles.Read',
            #'https://microsoft.sharepoint-df.com/MyFiles.Write',
            'https://graph.microsoft.com/Sites.ReadWrite.All',
            'https://graph.microsoft.com/offline_access',
            'basic',
            ]

    token_backend = O365.FileSystemTokenBackend(token_path='.', token_filename="my_token.txt")
    account = O365.Account(credentials, scopes=scopes, token_backend=token_backend)

    if not account.is_authenticated:
        account.authenticate()
        if not account.is_authenticated:
            log.fatal(f"Cannot authenticate account")
            sys.exit(1)


    # initialize volunteer connection api
    session = get_session(config)
    results = {
            'files': [],
            }

    if args.pull:
        xls0_buffer = read_active_positions(session, config)
        with open('test_active_positions.xls', 'w+b') as fh:
            fh.write(xls0_buffer)
    else:
        with open('test_active_positions.xls', 'rb') as fh:
            xls0_buffer = fh.read()
    process_active_positions(results, xls0_buffer)

    if args.pull:
        xls0_buffer = read_all_assignments(session, config)
        xls1_buffer = read_current_assignments(session, config)
        with open('test_all_assignments.xls', 'w+b') as fh:
            fh.write(xls0_buffer)
        with open('test_current_assignments.xls', 'w+b') as fh:
            fh.write(xls1_buffer)

    else:
        with open('test_all_assignments.xls', 'rb') as fh:
            xls0_buffer = fh.read()
        with open('test_current_assignments.xls', 'rb') as fh:
            xls1_buffer = fh.read()
    process_all_assignments(results, config, xls0_buffer, xls1_buffer)

    if args.pull:
        xls0_buffer = read_responder_availability(session, config)
        xls1_buffer = read_match_open(session, config)
        with open('test_avail.xls', 'w+b') as fh:
            fh.write(xls0_buffer)
        with open('test_match_open.xls', 'w+b') as fh:
            fh.write(xls1_buffer)
    else:
        with open('test_avail.xls', 'rb') as fh:
            xls0_buffer = fh.read()
        with open('test_match_open.xls', 'rb') as fh:
            xls1_buffer = fh.read()
    process_availability(results, xls0_buffer, xls1_buffer)

    if not args.post:
        return

    message = O365.Message(parent=account, main_resource=config.MESSAGE_FROM)

    #message.bcc.add(['neil@askneil.com'])
    message.to.add(['neil@askneil.com'])

    if args.post:
        #message.bcc.add(['peggy.pinarbasi@redcross.org'])
        message.to.add(['peggy.pinarbasi@redcross.org'])
        pass

    message.body = \
f"""<html>
<head>
<meta http-equiv="Content-type" content="text/html" charset="UTF8" />
</head>
<body>

<H1>Daily availability reports</H1>

<p>Hello Everyone.  Welcome to the new automated availability reports system.</p>

<p>Here are today's availability reports.</p>

<!--
<p>Summary information:<p>
<ul>
    <li><b>Metrics</b>
        <ul>
            <li>metric 1</li>
        </ul>
    </li>
</ul>

<p>If you want to be removed from the list or think something could be improved in these reports: send an email to <a href='mailto:DR534-21-Staffing-Reports@AmericanRedCross.onmicrosoft.com'>DR534-21-Staffing-Reports@AmericanRedCross.onmicrosoft.com</a>.</p>
-->

<p>These reports were run at { TIMESTAMP }.</p>

</body>
</html>
"""

    message.subject = f"{ config.MESSAGE_SUBJECT } { TIMESTAMP }"
    message.attachments.add(results['files'])
    message.send(save_to_sent_folder=True)


    # clean up after ourselves
    for file in results['files']:
        os.remove(file)

    return



ORDINAL_1900_01_01 = datetime.datetime(1900, 1, 1).toordinal()
TODAY = datetime.date.today()
TIMESTAMP = datetime.datetime.now().strftime('%Y-%m-%d %H%M')
LEFT_ALIGN = openpyxl.styles.Alignment(horizontal='left')


def process_availability(results, contents_responder_availability, contents_match_open):

    title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
    def pre_fixup(in_ws, out_ws, title):
        nonlocal title_font

        # copy the title values
        out_ws['A1'] = "NCCR Responder Availability"

        out_ws['a1'].font = title_font

    row_gap_type = None
    row_mem_num = None
    row_mem_name = None
    re_gap_type = re.compile(r'^(.*) (.*)$')


    def post_fixup(out_ws):
        out_ws["A2"] = 'GAP'


    def row_filter_avail(row, column_dict):
        """ do extra work to split GAP from Type """
        nonlocal row_gap_type

        # split the gap and type
        gap_col_index = column_dict['GAP and Type']
        gap = row[gap_col_index]
        match = re_gap_type.fullmatch(gap)
        if match != None:
            row[gap_col_index] = match.group(1)
            row_gap_type = match.group(2)
            if row_gap_type == '(none)':
                row_gap_type = ''
        else:
            row_gap_type = ''

        return row_filter_open(row, column_dict)


    def row_filter_open(row, column_dict):
        nonlocal row_mem_num
        nonlocal row_mem_name

        # stash the member num for county lookup
        row_mem_num = row[column_dict['Member Num']]
        row_mem_name = row[column_dict['Name']]


        # don't add folks who are restricted
        restricted = row[column_dict[' Rstr']]
        return restricted == ''
        
    def fill_gap_type(cell):
        nonlocal row_gap_type
        if row_gap_type != None:
            cell.value = row_gap_type


    if 'county_lookup' in results:
        county_lookup = results['county_lookup']
    else:
        county_lookup = {}

    def fill_county(cell):
        nonlocal county_lookup
        nonlocal row_mem_num
        nonlocal row_mem_name

        if row_mem_num != None and row_mem_num in county_lookup:
            county = county_lookup[row_mem_num]['county']
            #log.debug(f"looked up member { row_mem_num } / { row_mem_name } county { county } / { county_lookup[row_mem_num] }")
        else:
            log.debug(f"looked up member { row_mem_num } / { row_mem_name }: county is not in dict")
            county = 'unknown'

        cell.value = county

    re_date_convert = re.compile(r'(\d\d?)/(\d\d?)/(\d\d)')
    def fill_date_convert(cell):
        val = cell.value
        if isinstance(val, str):
            #log.debug(f"looking at val '{ val }'")
            match = re_date_convert.match(val)
            if match != None:
                dt = datetime.datetime(2000 + int(match.group(3)), int(match.group(1)), int(match.group(2)))
                #log.debug(f"converted date from '{ val }' to '{ val2 }' / { dt.strftime('%Y-%m-%d') }")
                cell.value = dt
                cell.number_format = 'yyyy-mm-dd'

    red_font = openpyxl.styles.Font(color='FFFF0000')
    def column_color(cell, in_color_index):
        """ look for in_color_index of 2 (red) and color output cell to match """

        if in_color_index == 2:
            #log.debug(f"saw red input cell for output { cell.coordinate }")
            cell.font = red_font

    def hyperlink_convert(cell, url):
        """ convert a hyperlink to output """
        #log.debug(f"saw hyperlink for output cell { cell.coordinate }")
        cell.hyperlink = url
        cell.style = 'Hyperlink'

    params = {
            'sheet_name':  'Availability',
            'out_file_name': f'NCCR Availability { TIMESTAMP }.xlsx',
            'table_name': 'Availability',
            'in_starting_row': 3,
            'out_starting_row': 2,
            'column_formats': {
                    'Avail Conf': 'yyyy-mm-dd',
                    },
            'column_widths': {
                    'GAP and Type': 12,
                    'Gap Type': 12,
                    'Avail Start': 12,
                    'Avail End': 12,
                    'Geo Avail': 14,
                    'Mem Av': 10,
                    'Avail Conf': 12,
                    'Name': 25,
                    'Preferred Name': 10,
                    'Member Num': 10,
                    'Release': 12,
                    'County': 13,
                    'City': 18,
                    'Email': 27,
                    'Home Phone': 13,
                    'Cell Phone': 13,
                    ' Qualifications': 50,
                    ' Languages': 20,
                    },
            'column_alignments': {
                    'Avail Start': LEFT_ALIGN,
                    'Avail End': LEFT_ALIGN,
                    'Avail Conf': LEFT_ALIGN,
                    'Release': LEFT_ALIGN,
                    },
            'column_fills': {
                    'Gap Type': fill_gap_type,
                    'County': fill_county,
                    'Avail Start': fill_date_convert,
                    'Avail End': fill_date_convert,
                    'Release': fill_date_convert,
                    },
            'pre_fixup': lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, "Disaster Responder Availability"),
            'post_fixup': post_fixup,
            'row_filter': row_filter_avail,
            'insert_columns': {
                    'Gap Type': 'Avail Start',
                    'County': 'City',
                    },
            'freeze_panes': 'B3',
            'column_colors': {
                    'Release': column_color,
                    'Avail Start': column_color,
                    'Avail End': column_color,
                    },
            'hyperlink_convert': {
                    'Name': hyperlink_convert,
                    },
            }

    results['files'].append(params['out_file_name'])
    out_wb = process_common(contents_responder_availability, params)

    params2 = {
            'sheet_name': 'Open Positions',
            'out_file_name': f'NCCR Availability { TIMESTAMP }.xlsx',
            'table_name': 'Open',
            'in_starting_row': 3,
            'out_starting_row': 2,
            'pre_fixup': lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, "Availability for Open Positions"),
            'freeze_panes': 'B3',
            'row_filter': row_filter_open,
            'column_fills': {
                    'County': fill_county,
                    'Avail Start': fill_date_convert,
                    'Avail End': fill_date_convert,
                    'Release': fill_date_convert,
                    },
            'column_widths': {
                    'DR name': 25,
                    'Req': 4,
                    'Open': 4,
                    'Notes': 20,
                    'Proximity':16,
                    'GAP': 24,
                    'Name': 25,
                    'Avail Start': 12,
                    'Avail End': 12,
                    'County': 13,
                    'Geo Avail': 14,
                    'POSITION qualification1': 24,
                    ' WORKER qualifications': 30,
                    'Member Num': 10,
                    'Release': 12,
                    'POSITION language': 20,
                    'WORKER languages': 40,
                    'Email': 27,
                    'Home Phone': 13,
                    'Cell Phone': 13,
                    'Reporting Location (if not virtual)': 40,
                    'DRO Hardship Codes': 40,
                    },
            'column_colors': {
                    'Release': column_color,
                    'Avail Start': column_color,
                    'Avail End': column_color,
                    'Proximity': column_color,
                    },
            'hyperlink_convert': {
                    'Name': hyperlink_convert,
                    },
            'insert_columns': {
                    'County': 'Geo Avail',
                    },
            }

    out_wb = process_common(contents_match_open, params2, out_wb=out_wb)



def process_all_assignments(results, config, contents_all_assignments, contents_current_assignments):
    """ process the all assignments spreadsheet """

    fill_today = openpyxl.styles.PatternFill(fgColor='C9E2B8', fill_type='solid')
    fill_tomorrow = openpyxl.styles.PatternFill(fgColor='9BC2E6', fill_type='solid')
    fill_past = openpyxl.styles.PatternFill(fgColor='FFDB69', fill_type='solid')

    results['arrive_today'] = 0
    results['arrive_tomorrow'] = 0

    last_row = False
    dro_name = None
    dro_number = None
    member_number = None
    member_name = None

    dro_name_split = re.compile(r'(\d+-\d+) (.*)')
    def parse_dro(number, name):
        """ the current sheet encodes the dro name in the dro number field; split it out """
        nonlocal dro_name_split

        if name == '':

            match = dro_name_split.match(number)
            if match != None:
                number = match.group(1)
                name = match.group(2)

        return (number, name)

    def pre_fixup(in_ws, out_ws, a1_value):
        nonlocal dro_name
        nonlocal dro_number

        # copy the title values
        out_ws['A1'] = a1_value

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['a1'].font = title_font

        dro_number = in_ws.cell_value(4,0)
        dro_name = in_ws.cell_value(4,1)

    def row_filter(row, column_dict, pass_type):
        """ filter out all title rows but the first one """
        nonlocal last_row
        nonlocal dro_name
        nonlocal dro_number
        nonlocal member_number
        nonlocal member_name

        if pass_type == 'all':
            last_row_val = 'People Assigned by DRO'
            mem_num_col_name = 'Mem#'
            col0_name = 'Mem#'
            col1_name = 'Name'
            col2_name = 'Chapter'
        elif pass_type == 'current':
            last_row_val = 'Region:'
            mem_num_col_name = 'Mem #'
            col0_name = 'Region'
            col1_name = 'Mem #'
            col2_name = 'Name'
        else:
            raise Exception(f"unknown pass_type '{ pass_type }'")

        if last_row:
            return False

        if row[0] == last_row_val:
            last_row = True
            return False

        mem_num = row[column_dict[mem_num_col_name]]
        member_name = row[column_dict['Name']]
        col0 = row[column_dict[col0_name]]
        col1 = row[column_dict[col1_name]]
        col2 = row[column_dict[col2_name]]
        #log.debug(f"row_filter: col0 '{ col0 }' col1 '{ col1 }' col2 '{ col2 }'")
        
        # filter out all other title rows
        if col2 == col2_name:
            return False

        if col0 == '' and col1 == '':
            # blank line
            return False

        if col2 == '':
            dro_number, dro_name = parse_dro(col0, col1)
            #log.debug(f"Capturing dro_number '{ dro_number }' dro_name '{ dro_name }'")
            return False

        # just a normal row: process it
        if isinstance(mem_num, (int, float)):
            member_number = int(mem_num)
            #log.debug(f"stashing member_number { member_number }")
        return True

    def fill_dro_name(cell):
        nonlocal dro_name
        cell.value = dro_name

    def fill_dro_number(cell):
        nonlocal dro_number
        cell.value = dro_number

    if 'county_lookup' in results:
        county_lookup = results['county_lookup']
    else:
        county_lookup = {}

    def fill_county(cell):
        nonlocal county_lookup
        nonlocal member_number

        if member_number != None and member_number in county_lookup:
            county = county_lookup[member_number]['county']
            #log.debug(f"looked up member { member_number } / { member_name } county { county } / { county_lookup[member_number] }")
        else:
            log.debug(f"looked up member { member_number } / { member_name }: county is not in dict")
            county = 'unknown'

        cell.value = county

    params = {
            'sheet_name': 'All Assignments by DR',
            'out_file_name': f'NCCR DR Assignments { TIMESTAMP }.xlsx',
            'table_name': 'AllAssignments',
            'in_starting_row': 5,
            'out_starting_row': 2,
            'column_formats': {
                    'Assign': 'yyyy-mm-dd',
                    'Release': 'yyyy-mm-dd',
                    },
            'column_widths': {
                    'Mem#': 10,
                    'Name': 25,
                    'DR Name': 26,
                    'DR Number': 10,
                    'County': 15,
                    'Last Action': 12,
                    'GAP': 13,
                    'Assign': 14,
                    'Release': 14,
                    'Category': 12,
                    'Cell phone': 13,
                    'Home phone': 13,
                    },
            'column_alignments': {
                    'Assign': LEFT_ALIGN,
                    'Release': LEFT_ALIGN,
                    },
            'column_fills': {
                    'DR Number': fill_dro_number,
                    'DR Name': fill_dro_name,
                    'County': fill_county,
                    },
            'pre_fixup': lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, f"All DR Assignments by DR in the last { config.ASSIGNMENT_DAYS } days"),
            'row_filter': lambda row, col_dict: row_filter(row, col_dict, 'all'),
            'insert_columns': {
                'DR Name': 'Chapter',
                'DR Number': 'DR Name',
                'County': 'Chapter',
                },
            'freeze_panes': 'C3',
            }

    results['files'].append(params['out_file_name'])
    out_wb = process_common(contents_all_assignments, params)

    params['sheet_name'] = 'Current Assignments'
    params['table_name'] = 'CurrentAssignments'
    params['column_formats'] = {
            'Assigned': 'yyyy-mm-dd',
            }
    params['column_alignments'] = {
                'Assigned': LEFT_ALIGN,
            }
    params['insert_columns'] = {
            'DR Name': 'Home City',
            'DR Number': 'DR Name',
            'County': 'Home City',
            }
    params['column_widths'] = {
                'Mem #': 10,
                'Name': 25,
                'DR Number': 10,
                'DR Name': 26,
                'County': 15,
                'Home City': 15,
                'Assigned': 14,
                'GAP': 12,
                'Category': 12,
                'Email': 30,
                'Cell phone': 13,
                'Home phone': 13,
            }
    params['pre_fixup'] =  lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, 'Currently assigned to a DR')
    params['row_filter'] =  lambda row, col_dict: row_filter(row, col_dict, 'current')
    params['freeze_panes'] = 'D3'

    last_row = False
    out_wb = process_common(contents_current_assignments, params, out_wb=out_wb)



def process_active_positions(results, contents):
    """ scan the contents xls file.  Pick out county.  index by member id """

    def get_row_value(row, column_map, row_title):
        # the row_title should always be in the column map...
        column = column_map[row_title]

        value = row[column]
        return value

    in_wb = xlrd.open_workbook(file_contents=contents)
    in_ws = in_wb.sheet_by_index(0)

    in_starting_row = 5
    delete_columns = []
    in_column_map, out_column_map = process_title_row(in_ws.row_values(in_starting_row), delete_columns)

    #log.debug(f"in_column_map: { in_column_map }")

    num_rows = in_ws.nrows - in_starting_row -1
    z_county_re = re.compile(r'Z County: (.*)')
    county_re = re.compile(r' County$')

    log.debug(f"num_rows { num_rows }")
    last_member_num = None
    z_county = None
    start_row = None
    result_dict = {}
    for index in range(num_rows +1):
        # read the row from the sheet

        if index >= num_rows:
            # we're past the last row
            member_num = None
        else:
            #log.debug(f"num_rows { num_rows } index { index } in_starting_row { in_starting_row }")
            in_row = in_ws.row_values(index + in_starting_row + 1)
            in_val = get_row_value(in_row, in_column_map, 'Member #')
            if in_val != '':
                member_num = int(in_val)
            else:
                member_num = 0

        if member_num != last_member_num:
            # this row is a new person; remember the id


            if last_member_num != None:
                # we're switching people

                if z_county != None:
                    county = z_county

                member_entry = {
                        'member_num': last_member_num,
                        'name': name,
                        'county': county,
                        'z-county': z_county,
                        'row-num': start_row,
                        }
                if county != '':
                    #log.debug(f"adding entry: member_num { member_num} to { member_entry }")
                    result_dict[last_member_num] = member_entry
                else:
                    #log.debug(f"empty county member_num { member_num} to { member_entry }")
                    pass

            # read out the cells we care about
            name = get_row_value(in_row, in_column_map, 'Account Name (hyperlink)')
            position = get_row_value(in_row, in_column_map, 'Position Name')
            county = get_row_value(in_row, in_column_map, 'County of Residence')
            county = county_re.sub('', county)

            z_county = None
            last_member_num = member_num
            start_row = index + in_starting_row + 2
        else:
            pass

        z_county_match = z_county_re.match(position)
        if z_county_match != None:
            z_county = z_county_match.group(1)

        #if index > 100:
        #    break

    #log.debug(f"result_dict: { pprint.pformat(result_dict, indent=2) }")
    results['county_lookup'] = result_dict



def process_common(contents, params, out_wb=None):
    """ common code to process all sheets """
    
    in_starting_row = params['in_starting_row']
    out_starting_row = params['out_starting_row']

    in_wb = xlrd.open_workbook(file_contents=contents, formatting_info=True)
    in_ws = in_wb.sheet_by_index(0)

    if out_wb == None:
        out_wb = openpyxl.Workbook()

    out_ws = out_wb.create_sheet(title=params['sheet_name'])

    # do some ws dependent preliminary initialization
    if 'pre_fixup' in params:
        params['pre_fixup'](in_ws, out_ws)

    # deal with the title row
    if 'delete_columns' in params:
        delete_columns = params['delete_columns']
    else:
        delete_columns = []

    # out_column_map maps column names to output columns, with columns to be ignored removed.
    # This could also reorder columns, but we're not using that functionality now.
    in_column_map, out_column_map = process_title_row(in_ws.row_values(in_starting_row), delete_columns)
    #log.debug(f"in_column_map { in_column_map }")

    if 'insert_columns' in params:
        for insert_col_name, insert_after in params['insert_columns'].items():
            # get the index of the existing column
            #log.debug(f"out_column_map before inserts: { out_column_map }")
            #log.debug(f"insert_col_name { insert_col_name } insert_after { insert_after }")
            insert_col_num = out_column_map[insert_after]
            new_out_map = {}
            for col_name, col_index in out_column_map.items():
                if col_index >= insert_col_num:
                    col_index += 1
                new_out_map[col_name] = col_index
            out_column_map = new_out_map
            out_column_map[insert_col_name] = insert_col_num
        #log.debug(f"out_column_map after inserts: { out_column_map }")


    col_format = {}
    if 'column_formats' in params:
        for col_name, cell_format in params['column_formats'].items():
            col_format[col_name] = cell_format

    if 'column_widths' in params:
        for col_name, col_width in params['column_widths'].items():
            col_index = out_column_map[col_name]
            col_letter = openpyxl.utils.cell.get_column_letter(col_index)
            out_ws.column_dimensions[col_letter].width = col_width

    col_alignment = {}
    if 'column_alignments' in params:
        for col_name, alignment in params['column_alignments'].items():
            col_alignment[col_name] = alignment

    col_fills = {}
    if 'column_fills' in params:
        for col_name, func in params['column_fills'].items():
            col_fills[col_name] = func

    col_colors = {}
    if 'column_colors' in params:
        for col_name, func in params['column_colors'].items():
            col_colors[col_name] = func

    col_hyperlink = {}
    if 'hyperlink_convert' in params:
        for col_name, func in params['hyperlink_convert'].items():
            col_hyperlink[col_name] = func

    # now deal with the body of the message
    num_rows = in_ws.nrows - in_starting_row
    row_filter = lambda x, y: True
    if 'row_filter' in params:
        log.debug("row filter is set")
        row_filter = params['row_filter']

    log.debug(f"num_rows { num_rows }")
    out_index = out_starting_row -1
    seen_xf = {}
    seen_font = {}
    for index in range(num_rows):
        in_row_index = index + in_starting_row
        in_row = in_ws.row_values(in_row_index)

        # allow us to ignore rows
        if index > 0 and not row_filter(in_row, in_column_map):
            #log.debug(f"ignoring row { index }")
            continue

        out_index += 1

        for col_name, out_col in out_column_map.items():
            if col_name in in_column_map:
                in_col_index = in_column_map[col_name]
                in_value = in_row[in_col_index]
            else:
                in_col_index = None
                in_value = ''

            if index == 0:
                in_value = col_name
            cell = out_ws.cell(row=out_index, column=out_col, value=in_value)

            if index > 0:
                if col_name in col_format:
                    # handle special column formats
                    cell.number_format = col_format[col_name]

                if col_name in col_fills:
                    col_fills[col_name](cell)

                if col_name in col_colors and in_col_index != None:
                    xf_index = in_ws.cell_xf_index(in_row_index, in_col_index)

                    if xf_index not in seen_xf:
                        xf = in_wb.xf_list[xf_index]
                        #log.debug(f"col_colors: cell row { in_row_index } col { in_col_index } xf_index { xf_index } xf\n{ xf.dump(indent=2) }\n\n")

                        font_index = xf.font_index
                        font = in_wb.font_list[font_index]
                        colour_index = font.colour_index

                        seen_xf[xf_index] = colour_index
                        #log.debug(f"new colour: cell row { in_row_index } col { in_col_index } font_index { font_index } colour { colour_index }")

                    col_colors[col_name](cell, seen_xf[xf_index])

                if col_name in col_hyperlink and in_col_index != None:
                    hyperlink = in_ws.hyperlink_map.get((in_row_index, in_col_index))
                    if hyperlink != None:
                        if hyperlink.type == 'url':
                            #log.debug(f"in cell { in_row_index },{in_col_index} hyperlink { hyperlink }")
                            #hyperlink.dump(indent=2)
                            col_hyperlink[col_name](cell, hyperlink.url_or_path)


            if col_name in col_alignment:
                cell.alignment = col_alignment[col_name]

    # do some ws dependent table fixup after the copy
    if 'post_fixup' in params:
        params['post_fixup'](out_ws)


    # now make a table of the data
    start_col = 'A'
    end_col = openpyxl.utils.cell.get_column_letter(len(out_column_map))
    table_ref = f"{start_col}{out_starting_row}:{end_col}{out_starting_row + num_rows -1}"
    #log.debug(f"table_ref '{table_ref}'")
    table = openpyxl.worksheet.table.Table(displayName=params['table_name'], ref=table_ref)
    out_ws.add_table(table)

    if 'freeze_panes' in params:
        freeze_panes = params['freeze_panes']
        log.debug(f"setting freeze_panes to cell '{ freeze_panes }'")
        out_ws.freeze_panes = out_ws[freeze_panes]

    default_sheet_name = 'Sheet'
    if default_sheet_name in out_wb:
        del out_wb[default_sheet_name]

    out_wb.save(params['out_file_name'])

    return out_wb


def process_title_row(row, delete_columns):
    """ process an xlrd title row, returning a map of column names to column indexes (origin zero) """
    in_column_map = {}
    out_column_map = {}
    out_col = 0
    for col in range(len(row)):
        value = row[col]
        if value is not None and value != '':

            if value not in delete_columns:
                in_column_map[value] = col
                out_col += 1
                out_column_map[value] = out_col

    return in_column_map, out_column_map

def read_match_open(session, config):
    """ run the Disaster Responder Availability Match Open Positions report """

    params0 = {
            'query_id': '1972837',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': '155',

            'prompt2': [
                'Unknown__,__',
                'National__,__',
                'Division__,__',
                'Region__,__',
                'State__,__',
                'Virtual__,__',
                ],

            'prompt3': 'Available today or within a week',
            'prompt4': 'No',
            'prompt5': 'No',
            'prompt6': 'ALL',
            'prompt7': 'Name',

            'output_format': 'xls',
            'run': 'Run',
            }

    # give an extra-long timeout, since the report is slow
    params1 = convert_params(params0)

    log.debug(f"params1:\n{ pprint.pformat(params1, indent=2) }")

    return read_common(session, config, params0, params1, timeout=config.WEB_TIMEOUT + 60)

def read_responder_availability(session, config):
    """ run the Distaster Responder Availability by GAP report """

    params0 = {
            'query_id': '1803665',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': '155',

            'prompt2': [
                'DST__,__',
                'ER__,__',
                'FIN__,__',
                'IDC__,__',
                'IP__,__',
                'LOG__,__',
                'MC__,__',
                'OM__,__',
                'REC__,__',
                'RES__,__',
                'SS__,__',
                ],

            'prompt3': [
                '__,__',
                '-1__,__',
                'APAT__,__',
                'CEP__,__',
                'CM__,__',
                'CMR__,__',
                'CO__,__',
                'CRP__,__',
                'CS__,__',
                'DA__,__',
                'DAT__,__',
                'DEBV__,__',
                'DES__,__',
                'DHS__,__',
                'DI__,__',
                'DMH__,__',
                'DSC__,__',
                'ER__,__',
                'FAC__,__',
                'FF__,__',
                'FIN__,__',
                'FR__,__',
                'FSI__,__',
                'GEN__,__',
                'GO__,__',
                'ID__,__',
                'IKD__,__',
                'IP__,__',
                'LCV__,__',
                'LOG__,__',
                'LSAP__,__',
                'NT__,__',
                'OPS__,__',
                'PA__,__',
               'PRO__,__',
                'REU__,__',
                'SH__,__',
                'SPS__,__',
                'SR__,__',
                'SU__,__',
                'SUP__,__',
                'SYS__,__',
                'TR__,__',
                'TRA__,__',
                'WF__,__',
                'WHS__,__',
                ],

            'prompt4': [
                'AC__,__',
                'AD__,__',
                'CH__,__',
                'COS__,__',
                'DD__,__',
                'DIR__,__',
                'EOL__,__',
                'MD__,__',
                'MN__,__',
                'RCCO__,__',
                'SA__,__',
                'SD__,__',
                'STA__,__',
                'SV__,__',
                ],

            'prompt5': [
                'Unknown__,__',
                'National__,__',
                'Division__,__',
                'Region__,__',
                'State__,__',
                'Virtual__,__',
                ],
            'prompt6': 'Available today or within a week',
            'output_format': 'xls',
            'run': 'Run',
            }

    # give an extra-long timeout, since the report is slow
    params1 = convert_params(params0)

    log.debug(f"params1:\n{ pprint.pformat(params1, indent=2) }")

    return read_common(session, config, params0, params1, timeout=config.WEB_TIMEOUT + 60)


def read_all_assignments(session, config):
    """ run the All Assignments By DR and/or Date Range report """

    dt_days = datetime.datetime.now() - datetime.timedelta(days=config.ASSIGNMENT_DAYS)
    s_days = dt_days.strftime('%Y-%m-%d')

    #log.debug(f"s_days = '{ s_days }'")


    params0 = {
            'query_id': '887302',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': '155',
            'prompt2': 'Yes',
            'prompt3': s_days,
            'prompt4': '',
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = convert_params(params0)

    #log.debug(f"params1:\n{ pprint.pformat(params1, indent=2) }")

    return read_common(session, config, params0, params1)


def read_current_assignments(session, config):
    """ run the Disaster Responders Currently Assigned - Region report """

    params0 = {
            'query_id': '38613',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': '155',
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = convert_params(params0)

    return read_common(session, config, params0, params1)


def read_active_positions(session, config):


    params0 = {
            'query_id': '1394341',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': '155',
            'prompt2': [ 'Employee__,__', 'Volunteer__,__', ],
            'prompt3': [ '369__,__', '366__,__', '1099__,__', '375__,__', '367__,__', ],
            'prompt4': [ 'Current__,__', 'Ended__,__', ],
            'prompt5': [ '-2__,__', '2766__,__', ],
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = convert_params(params0)

    log.debug(f"params1:\n{ pprint.pformat(params1, indent=2) }")

    return read_common(session, config, params0, params1)


def convert_date(dt):
    """ special case for formatting date parameter for param1 """
    return dt.strftime('Date(%Y,%m,%d)')

RE_PARAM_SUFFIX = re.compile(r'__,__$')
RE_DATE = re.compile(r'^(\d{4})-(\d{2})-(\d{2})$')
def convert_params(params0):

    def convert(param, index):
        if index == 0:
            return param

        if isinstance(param, (str)):

            if RE_PARAM_SUFFIX.search(param):
                return f"['{ RE_PARAM_SUFFIX.sub('', param) }']"

            if RE_DATE.search(param):
                return RE_DATE.sub(r'Date(\1,\2,\3)', param)

            # didn't match either of the patterns; just return the string
            return param
        elif isinstance(param, list):
            arg = ",".join( f"'{ RE_PARAM_SUFFIX.sub('', p) }'" for p in param )
            return f"[{ arg }]"
        else:
            raise Exception(f"unexpected parameter type { type(param) }")
    
    params1 = {}

    params1['nd'] = 'clearreports_auth'
    params1['init'] = params0['output_format']
    params1['query_id'] = params0['query_id']

    param_num = 0
    params0_name = f"prompt{ param_num + 1 }"
    params1_name = f"prompt{ param_num }"
    params1[params1_name] = params0[params0_name]

    while True:


        param_num += 1
        params0_name = f"prompt{ param_num + 1 }"
        params1_name = f"prompt{ param_num }"


        if params0_name not in params0:
            break

        params1[params1_name] = convert(params0[params0_name], param_num)

    return params1


def read_common(session, config, params0, params1, timeout=None):

    if timeout == None:
        timeout = config.WEB_TIMEOUT


    url = "https://volunteerconnection.redcross.org/"

    headers = {
            #'accept': 'application/json, text/javascript, */*; q=0.01',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'DNT': '1',
            'Host': 'volunteerconnection.redcross.org',
            'Origin': 'https://volunteerconnection.redcross.org',
            'Referer': 'https://volunteerconnection.redcross.org/',
            #'Sec-Fetch-Dest': 'empty',
            #'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'sec-gpc': '1',
            'User-Agent': '.Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            }
    response = session.post(url, data=params0, headers=headers, timeout=timeout)
    response.raise_for_status()


    response = session.post(url, data=params1, headers=headers, timeout=timeout)
    response.raise_for_status()

    #log.debug(f"response.content { response.content }")
    anchor = response.html.find('a', first=True)
    href = anchor.attrs['href']

    url2 = url + href

    #log.debug(f"url2 { url2 }")
    response = session.get(url2, timeout=timeout)
    response.raise_for_status()

    log.debug(f"retrieved document.  size is { len(response.text) }, type is '{ response.headers.get('content-type') }'")
    #log.debug(f"document: { response.content }")

    return response.content

def init_config():
    class AttrDict(dict):
        def __init__(self, *args, **kwargs):
            super(AttrDict, self).__init__(*args, **kwargs)
            self.__dict__ = self


    config_dotenv = dotenv.dotenv_values(verbose=True)

    config = AttrDict()
    for item in dir(config_static):
        if not item.startswith('__'):
            config[item] = getattr(config_static, item)

    #log.debug(f"config after copy: { config.keys() }")

    for key, val in config_dotenv.items():
        config[key] = val

    return config

def parse_args():
    parser = argparse.ArgumentParser(
            description="process support for the regional bootcamp mission card system",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("--post", help="post to real recipients", action="store_true")
    parser.add_argument("--pull", help="read real data from VC", action="store_true")

    args = parser.parse_args()
    return args



if __name__ == "__main__":
    main()


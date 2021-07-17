#! /usr/bin/env python3

# gaptastic -- match open door positions to available responders

import argparse
import logging
import os
import os.path
import re
import datetime
import time
import pprint
import json
import io
import csv
import sys
import random

import requests
import requests_html
import dotenv
import xlrd
import openpyxl

from http.cookiejar import LWPCookieJar, Cookie


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import O365

import init_logging
from vc_session import get_session
log = logging.getLogger(__name__)
import config as config_static


def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    global config
    config = init_config()

    # initialize office 365 graph api
    credentials = (config.CLIENT_ID, config.CLIENT_SECRET)

    scopes = [
            'https://graph.microsoft.com/Files.ReadWrite.All',
            'https://graph.microsoft.com/Mail.Send',
            'https://graph.microsoft.com/Mail.Send.Shared',
            'https://graph.microsoft.com/offline_access',
            'https://graph.microsoft.com/User.Read',
            'https://graph.microsoft.com/User.ReadBasic.All',
            #'https://microsoft.sharepoint-df.com/AllSites.Read',
            #'https://microsoft.sharepoint-df.com/MyFiles.Read',
            #'https://microsoft.sharepoint-df.com/MyFiles.Write',
            'https://graph.microsoft.com/Sites.ReadWrite.All',
            'https://graph.microsoft.com/offline_access',
            'basic',
            ]

    token_backend = O365.FileSystemTokenBackend(token_path='.', token_filename="my_token.txt")
    account = O365.Account(credentials, scopes=scopes, token_backend=token_backend)

    if not account.is_authenticated:
        account.authenticate()
        if not account.is_authenticated:
            log.fatal(f"Cannot authenticate account")
            sys.exit(1)


    # initialize volunteer connection api
    session = get_session(config)
    results = {
            'files': [],
            }

    # for debugging new reports
    #with open("air_travel_roster.xls", 'wb') as file:
    #    data = read_air_travel_roster(session, config, False)
    #    file.write(data)
    #with open("air_travel_roster.xls", 'rb') as file:
    #    data = file.read()
    #process_air_travel_roster(results, data)
    #return

    #with open("open_requests.xls", 'wb') as fd:
    #    data = read_open_requests(session, config, False)
    #    fd.write(data)

    process_arrival_roster(results, read_arrival_roster(session, config, False))
    process_open_requests(results, read_open_requests(session, config, False))
    process_staff_roster(results, read_staff_roster(session, config, False))
    process_air_travel_roster(results, read_air_travel_roster(session, config, False))
    #process_shift_tool(results, read_shift_tool(session, config, False))

    mailbox = account.mailbox()
    message = mailbox.new_message()

    #attach0_body = "attachment body\n"
    #attach0 = O365.utils.attachment.BaseAttachment(
    #        parent=message,
    #        content=attach0_body,
    #        size=len(attach0_body),
    #        content_type="text/plain",
    #        name='attach0.txt')

    message.bcc.add([config.MAIL_BCC])

    if args.post:
        message.bcc.add([config.MAIL_ADDRESS])
        log.debug(f"--post arg seen; sending to { config.MAIL_ADDRESS } also")


    message.body = \
f"""<html>
<head>
<meta http-equiv="Content-type" content="text/html" charset="UTF8" />
</head>
<body>

<H1>DR Staff Reports</H1>

<p>Hello Everyone.  Welcome to the automated staffing reports system.</p>

<p>Here are the current staff reports.</p>

<p>Summary information:<p>
<ul>
    <li><b>Staff Counts</b>
        <ul>
            <li>{ results['staff_total'] } active responders assigned to the job (both checked in and due to arrive)</li>
            <!-- <li>{ results['staff_nccr'] } of those from NCCR</li> -->
            <li>{ results['arrive_today'] } on the arrival roster for today</li>
            <li>{ results['arrive_tomorrow'] } on the arrival roster for tomorrow</li>
            <li>{ results['staff_outprocessed'] } out-processed</li>
        </ul>
    </li>
    <li><b>Staff Requests</b>: { results['requests_requests'] } requests for { results['requests_open'] } Open Positions
</ul>


<p>
This message was sent to { config.MAIL_ADDRESS }.  You can see old reports in
<a href='{ config.MAIL_ARCHIVE }'> the list archive</a>
(if you have a redcross.org account...)
</p>

<p>
If you wish to be removed from the group or have more people added: email
<a href='mailto:{ config.MAIL_OWNER }'>{ config.MAIL_OWNER}</a>
</p>

<!--
<p>The <b>DRO Shift Tool Roster</b> has been added to give a picture of the DRO shifts from yesterday, as well registered shifts for today and tomorrow. These workers don’t show up on the regular staff roster, so if you need to get ahold of them, you will find their contact information in the report.</p>

<p>If you have a roster change to submit you can do so
<a href='https://volunteerconnection.redcross.org/?nd=vms_public_form&form_id=8562'>on this form</a></p>

<p>If you want to be removed from the list or think something could be improved in these reports: send an email to <a href='mailto:DR534-21-Staffing-Reports@AmericanRedCross.onmicrosoft.com'>DR534-21-Staffing-Reports@AmericanRedCross.onmicrosoft.com</a>.</p>
-->


<p>These reports were run at { TIMESTAMP }.</p>

</body>
</html>
"""

    message.subject = f"Staff Reports { TIMESTAMP }"
    message.attachments.add(results['files'])
    message.send(save_to_sent_folder=True)


    # clean up after ourselves
    if not args.keep_files:
        for file in results['files']:
            os.remove(file)

    return



ORDINAL_1900_01_01 = datetime.datetime(1900, 1, 1).toordinal()
TODAY = datetime.date.today()
TIMESTAMP = datetime.datetime.now().strftime('%Y-%m-%d %H%M')
LEFT_ALIGN = openpyxl.styles.Alignment(horizontal='left')

def process_air_travel_roster(results, contents):

    #fill_today = openpyxl.styles.PatternFill(fgColor='C9E2B8', fill_type='solid')
    #fill_tomorrow = openpyxl.styles.PatternFill(fgColor='9BC2E6', fill_type='solid')
    #fill_past = openpyxl.styles.PatternFill(fgColor='FFDB69', fill_type='solid')

    #results['arrive_today'] = 0
    #results['arrive_tomorrow'] = 0

    def pre_fixup(in_ws, out_ws, params):
        # copy the title values
        out_ws['A1'] = in_ws.cell_value(0,2)    # report name
        out_ws['D1'] = in_ws.cell_value(2,0)    # DR string
        out_ws['A2'] = in_ws.cell_value(0,11)    # Date time label (with Timezone)
        out_ws['D2'] = in_ws.cell_value(1,11) + in_ws.cell_value(1,14)   # Date Value + Time Value
        out_ws['D2'].number_format = 'yyyy-mm-dd HH:MM'
        out_ws['D2'].alignment = LEFT_ALIGN

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['a1'].font = out_ws['d1'].font = title_font

        #out_ws.cell(row=1, column=11, value='Arriving Today').fill = fill_today
        #out_ws.cell(row=2, column=11, value='Arriving Tomorrow').fill = fill_tomorrow
        #out_ws.cell(row=3, column=11, value='Past Due Date').fill = fill_past

        #in_starting_row = params['in_starting_row']

        #if in_ws.cell_value(in_starting_row, 0) != 'Name':
        #    if in_ws.cell_value(in_starting_row-1, 0) == 'Name':
        #        # vc has fewer header rows in an empty spreadsheet...  Sigh.
        #        params['in_starting_row'] = in_starting_row -1

    def filter_arrive_date(cell, today, fill_past, fill_today, fill_tomorrow):
        """ decide if there is a special fill to apply to the cell """
        if cell.value == "":
            return ""

        excel_date = cell.value
        dt = datetime.datetime.fromordinal(ORDINAL_1900_01_01 + int(excel_date) -2)
        date = dt.date()

        if date < today:
            cell.fill = fill_past
        elif date == today:
            cell.fill = fill_today
            results['arrive_today'] += 1
        elif date == datetime.timedelta(1) + today:
            cell.fill = fill_tomorrow
            results['arrive_tomorrow'] += 1

    params = {
            'sheet_name': 'Air Travel Roster',
            'out_file_name': f'{ config.DR_NAME} Air Travel Roster { TIMESTAMP }.xlsx',
            'table_name': 'AirTravel',
            'in_starting_row': 3,
            'out_starting_row': 3,
            'column_formats': {
                    'Arrival Date/Time': 'yyyy-mm-dd HH:MM',
                    'Flight Arrival Date/Time': 'yyyy-mm-dd HH:MM',
                    },
            'column_widths': {
                    'Mem#': 10,
                    'Name': 25,
                    'Gen': 4,
                    'Arrival Date/Time': 20,
                    'Arrival City': 20,
                    'Departure City': 20,
                    'GAP': 17,
                    'Airline': 17,
                    'Assign/CheckIn': 12,
                    'Cell Number': 13,
                    'Status': 19, 
                    'Region name': 32,
                    },
            'column_alignments': {
                    'Arrival Date/Time': LEFT_ALIGN,
                    },
            'column_fills': {
                    #'Arrive date': lambda cell: filter_arrive_date(cell, TODAY, fill_past, fill_today, fill_tomorrow),
                    },
            'pre_fixup': lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, params),
            }

    results['files'].append(params['out_file_name'])


    process_common(contents, params)

def process_arrival_roster(results, contents):

    fill_today = openpyxl.styles.PatternFill(fgColor='C9E2B8', fill_type='solid')
    fill_tomorrow = openpyxl.styles.PatternFill(fgColor='9BC2E6', fill_type='solid')
    fill_past = openpyxl.styles.PatternFill(fgColor='FFDB69', fill_type='solid')

    results['arrive_today'] = 0
    results['arrive_tomorrow'] = 0

    def pre_fixup(in_ws, out_ws, params):
        # copy the title values
        out_ws['A1'] = in_ws.cell_value(0,0)
        out_ws['A2'] = in_ws.cell_value(1,0)
        out_ws['A3'] = in_ws.cell_value(2,0)

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['k1'].font = out_ws['k2'].font = out_ws['k3'].font = title_font
        out_ws['a1'].font = out_ws['a2'].font = out_ws['a3'].font = title_font

        out_ws.cell(row=1, column=11, value='Arriving Today').fill = fill_today
        out_ws.cell(row=2, column=11, value='Arriving Tomorrow').fill = fill_tomorrow
        out_ws.cell(row=3, column=11, value='Past Due Date').fill = fill_past

        in_starting_row = params['in_starting_row']

        if in_ws.cell_value(in_starting_row, 0) != 'Name':
            if in_ws.cell_value(in_starting_row-1, 0) == 'Name':
                # vc has fewer header rows in an empty spreadsheet...  Sigh.
                params['in_starting_row'] = in_starting_row -1

    def filter_arrive_date(cell, today, fill_past, fill_today, fill_tomorrow):
        """ decide if there is a special fill to apply to the cell """
        if cell.value == "":
            return ""

        excel_date = cell.value
        dt = datetime.datetime.fromordinal(ORDINAL_1900_01_01 + int(excel_date) -2)
        date = dt.date()

        if date < today:
            cell.fill = fill_past
        elif date == today:
            cell.fill = fill_today
            results['arrive_today'] += 1
        elif date == datetime.timedelta(1) + today:
            cell.fill = fill_tomorrow
            results['arrive_tomorrow'] += 1

    params = {
            'sheet_name': 'Arrival Roster',
            'out_file_name': f'{ config.DR_NAME } Arrival Roster { TIMESTAMP }.xlsx',
            'table_name': 'Arrival',
            'in_starting_row': 5,
            'out_starting_row': 4,
            'column_formats': {
                    'Arrive date': 'yyyy-mm-dd',
                    'Flight Arrival Date/Time': 'yyyy-mm-dd HH:MM',
                    },
            'column_widths': {
                    'Name': 25,
                    'Flight Arrival Date/Time': 25,
                    'Flight City': 18,
                    'District': 18,
                    'GAP': 12,
                    'Arrive date': 14,
                    'Reporting/Work Location': 22,
                    'Email': 25,
                    'Cell phone': 13,
                    'Home phone': 13,
                    'Work phone': 13,
                    },
            'column_alignments': {
                    'Arrive date': LEFT_ALIGN,
                    'Flight Arrival Date/Time': LEFT_ALIGN,
                    },
            'column_fills': {
                    'Arrive date': lambda cell: filter_arrive_date(cell, TODAY, fill_past, fill_today, fill_tomorrow),
                    },
            'pre_fixup': lambda in_ws, out_ws: pre_fixup(in_ws, out_ws, params),
            }

    results['files'].append(params['out_file_name'])

    # wierd things happen if arrival roster is empty: the title row is one row before it should be

    process_common(contents, params)


def process_open_requests(results, contents):

    def pre_fixup(in_ws, out_ws):
        # copy the title values
        out_ws['A1'] = in_ws.cell_value(0,0)
        out_ws['E1'] = in_ws.cell_value(0,3)

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['A1'].font = out_ws['E1'].font = title_font

    results['requests_open'] = 0
    results['requests_requests'] = 0

    def filter_req(cell, results):
        results['requests_requests'] += 1


    def filter_open(cell, results):
        value = cell.value

        try:
            results['requests_open'] += int(value)
        except:
            pass


    params = {
            'sheet_name': 'Open Staff Requests',
            'out_file_name': f'{ config.DR_NAME } Open Staff Requests { TIMESTAMP }.xlsx',
            'table_name': 'OpenRequests',
            'in_starting_row': 1,
            'out_starting_row': 2,
            'column_formats': {
                    },
            'column_widths': {
                    'Proximity': 16,
                    'G/A/P': 12,
                    'Qual 1': 15,
                    'Location': 18,
                    'Supervisor': 24,
                    },
            'column_alignments': {
            #        'Arrive date': LEFT_ALIGN,
            #        'Flight Arrival Date/Time': LEFT_ALIGN,
                    },
            
            'column_fills': {
                    'Req':  lambda cell: filter_req(cell, results),
                    'Open': lambda cell: filter_open(cell, results),
                    },
            'pre_fixup': pre_fixup,
            #'post_fixup': post_fixup,
            }

    results['files'].append(params['out_file_name'])
    process_common(contents, params)



def process_staff_roster(results, contents):
    """ generate the staff roster spreadsheets """

    fill_remain = openpyxl.styles.PatternFill(fgColor='FFDB69', fill_type='solid')

    def pre_fixup(in_ws, out_ws):
        # copy the title values
        out_ws['A1'] = in_ws.cell_value(0,0)
        out_ws['A2'] = in_ws.cell_value(1,0)
        out_ws['A3'] = in_ws.cell_value(2,0)

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['A1'].font = out_ws['A2'].font = out_ws['A3'].font = title_font

    def row_filter(row, out_column_map):
        released = row[out_column_map['Released']]
        #log.debug(f"checking released: { released }")

        return released == ''

    def filter_days_remain(cell, today, fill_remain):

        value = cell.value
        if value != "" and value != 'n/a':
            value = cell.value = int(cell.value)
            if value <= 2:
                cell.fill = fill_remain

    def filter_on_job(cell):
        value = cell.value
        if value != "" and value != 'n/a':
            cell.value = int(cell.value)


    results['staff_total'] = 0
    results['staff_nccr'] = 0
    results['staff_outprocessed'] = 0

    def filter_region(cell, results):
        """ don't change anything; just count total and NCCR folks """

        NCCR = '05R28'
        value = cell.value

        results['staff_total'] += 1

        if value == NCCR:
            results['staff_nccr'] += 1

    def filter_outprocessed(cell, results):
        """ count total outprocessed people """
        results['staff_outprocessed'] += 1


    def post_fixup_staff(ws):

        cell = ws['A2']
        value = cell.value

        regex = r'\(\d+ '
        cell.value = re.sub(regex, f"({results['staff_total']} ", value)
        #log.debug(f"value '{value}' after '{cell.value}'")

    def post_fixup_outprocessed(ws):

        cell = ws['A2']
        value = cell.value

        regex = r'\(\d+ '
        cell.value = re.sub(regex, f"({results['staff_outprocessed']} ", value)
        #log.debug(f"value '{value}' after '{cell.value}'")

    params = {
            'sheet_name': 'Staff Roster',
            'out_file_name': f'{ config.DR_NAME } Staff Roster { TIMESTAMP }.xlsx',
            'table_name': 'StaffRoster',
            'in_starting_row': 5,
            'out_starting_row': 4,
            'column_formats': {
                    'Assigned': 'yyyy-mm-dd',
                    'Checked in': 'yyyy-mm-dd',
                    'Released': 'yyyy-mm-dd',
                    'Expect release': 'yyyy-mm-dd',
                    },
            'column_widths': {
                    'Name': 25,
                    'Assigned': 11,
                    'Checked in': 11,
                    'Released': 11,
                    'Current/Last Supervisor': 22,
                    'GAP(s)': 14,
                    'District': 16,
                    'Reporting/Work Location': 36,
                    'Location type': 16,
                    'Expect release': 11,
                    'Last action': 12,
                    'Current lodging': 20,
                    'Qualifications': 32,
                    'All GAPs': 32,
                    'Languages': 28,
                    'All Supervisors': 28,
                    'Evaluation status(es)': 28,
                    'COVID-19 issuer/ notes': 48,
                    'Email': 33,
                    'Cell phone': 14,
                    'Home phone': 14,
                    'Work phone': 14,
                    ' ZIP': 12,         # yes, there really is a space in the column title as generated by VC
                    },
            'column_alignments': {
                    'Assigned': LEFT_ALIGN,
                    'Checked in': LEFT_ALIGN,
                    'Released': LEFT_ALIGN,
                    'Expect release': LEFT_ALIGN,
                    },
            
            'column_fills': {
                    'DaysRemain': lambda cell: filter_days_remain(cell, TODAY, fill_remain),
                    'On Job': lambda cell: filter_on_job(cell),
                    'Region': lambda cell: filter_region(cell, results),
                    },
            'pre_fixup': pre_fixup,
            'post_fixup': post_fixup_staff,
            'row_filter': row_filter,
            }

    results['files'].append(params['out_file_name'])
    process_common(contents, params)

    params['sheet_name'] = 'Outprocessed'
    params['out_file_name'] = f'{ config.DR_NAME } Outprocessed Roster { TIMESTAMP }.xlsx'
    params['table_name'] = 'Outprocessed'
    params['row_filter'] = lambda row, out_column_map: not row_filter(row, out_column_map)
    params['column_fills']['Region'] = lambda cell: filter_outprocessed(cell, results)
    params['post_fixup'] = post_fixup_outprocessed

    results['files'].append(params['out_file_name'])
    process_common(contents, params)



def process_shift_tool(results, contents):
    """ prepare the dro shift tool spreadsheet """

    fill_today = openpyxl.styles.PatternFill(fgColor='C9E2B8', fill_type='solid')
    fill_tomorrow = openpyxl.styles.PatternFill(fgColor='9BC2E6', fill_type='solid')

    def filter_days(cell, today, fill_today, fill_tomorrow):
        """ decide if there is a special fill to apply to the cell """
        if cell.value == '':
            return
        excel_date = cell.value
        dt = datetime.datetime.fromordinal(ORDINAL_1900_01_01 + int(excel_date) -2)
        date = dt.date()

        if date == today:
            cell.fill = fill_today
        elif date == datetime.timedelta(1) + today:
            cell.fill = fill_tomorrow


    def pre_fixup(in_ws, out_ws):
        # copy the title values
        out_ws['A1'] = in_ws.cell_value(0,0)
        out_ws['A2'] = in_ws.cell_value(1,0)

        out_ws['F1'] = 'Shifts Scheduled Today'
        out_ws['F2'] = 'Shifts Scheduled Tomorrow'
        out_ws['F1'].fill = fill_today
        out_ws['F2'].fill = fill_tomorrow
        out_ws['G1'].fill = fill_today
        out_ws['G2'].fill = fill_tomorrow
        out_ws['H1'].fill = fill_today
        out_ws['H2'].fill = fill_tomorrow

        title_font = openpyxl.styles.Font(name='Arial', size=14, bold=True)
        out_ws['A1'].font = out_ws['A2'].font = title_font
        out_ws['F1'].font = out_ws['F2'].font = title_font


    params = {
            'sheet_name': 'DRO Shift Tool',
            'out_file_name': f'DRO Shift Tool { TIMESTAMP }.xlsx',
            'table_name': 'ShiftTool',
            'in_starting_row': 2,
            'out_starting_row': 3,
            'column_formats': {
                    'Start Date': 'yyyy-mm-dd',
                    'Start Time': 'hh:mm AM/PM',
                    },
            'column_widths': {
                    'Name': 36,
                    'Volunteer Status': 28,
                    'Email': 38,
                    'Phone Numbers': 36,
                    'Shift Name': 36,
                    'Start Date': 12,
                    'Start Time': 12,
                    'Shift Location': 48,
                    },
            'column_alignments': {
                    'Start Date': LEFT_ALIGN,
                    'Start Time': LEFT_ALIGN,
                    },
            'delete_columns': [
                'Account ID',
                'Address',
                'District (of shift)',
                'Attended/Sign In',
                'Type of ID Presented',
                ],
            
            'column_fills': {
                    'Start Date': lambda cell: filter_days(cell, TODAY, fill_today, fill_tomorrow),
                    },
            'pre_fixup': pre_fixup,
            #'post_fixup': post_fixup,
            #'row_filter': row_filter,
            }

    results['files'].append(params['out_file_name'])
    process_common(contents, params)


def process_common(contents, params):
    """ common code to process all sheets """
    

    in_wb = xlrd.open_workbook(file_contents=contents)
    in_ws = in_wb.sheet_by_index(0)

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.create_sheet(title=params['sheet_name'])

    # do some ws dependent preliminary initialization
    if 'pre_fixup' in params:
        params['pre_fixup'](in_ws, out_ws)

    in_starting_row = params['in_starting_row']
    out_starting_row = params['out_starting_row']

    # deal with the title row
    if 'delete_columns' in params:
        delete_columns = params['delete_columns']
    else:
        delete_columns = []

    # out_column_map maps column names to output columns, with columns to be ignored removed.
    # This could also reorder columns, but we're not using that functionality now.
    in_column_map, out_column_map = process_title_row(in_ws.row_values(in_starting_row), delete_columns)

    #log.debug(f"in_column_map { in_column_map }")
    #log.debug(f"out_column_map { out_column_map }")

    col_format = {}
    for col_name, cell_format in params['column_formats'].items():
        col_format[col_name] = cell_format

    for col_name, col_width in params['column_widths'].items():
        col_index = out_column_map[col_name]
        col_letter = openpyxl.utils.cell.get_column_letter(col_index)
        out_ws.column_dimensions[col_letter].width = col_width

    col_alignment = {}
    for col_name, alignment in params['column_alignments'].items():
        col_alignment[col_name] = alignment

    col_fills = {}
    for col_name, fill in params['column_fills'].items():
        col_fills[col_name] = fill

    # now deal with the body of the message
    num_rows = in_ws.nrows - in_starting_row
    row_filter = lambda x, y: True
    if 'row_filter' in params:
        row_filter = params['row_filter']

    #log.debug(f"num_rows { num_rows }")
    out_index = out_starting_row -1
    for index in range(num_rows):
        in_row = in_ws.row_values(index + in_starting_row)

        # allow us to ignore rows
        if index > 0 and not row_filter(in_row, in_column_map):
            continue

        out_index += 1

        for col_name, out_col in out_column_map.items():
            in_col = in_column_map[col_name]
            cell = out_ws.cell(row=out_index, column=out_col, value=in_row[in_col])

            if index > 0:
                if col_name in col_format:
                    # handle special column formats
                    cell.number_format = col_format[col_name]

                if col_name in col_fills:
                    col_fills[col_name](cell)

            if col_name in col_alignment:
                cell.alignment = col_alignment[col_name]

    # do some ws dependent table fixup after the copy
    if 'post_fixup' in params:
        params['post_fixup'](out_ws)


    # now make a table of the data
    start_col = 'A'
    end_col = openpyxl.utils.cell.get_column_letter(len(out_column_map))
    table_ref = f"{start_col}{out_starting_row}:{end_col}{out_starting_row + num_rows -1}"
    #log.debug(f"table_ref '{table_ref}'")
    table = openpyxl.worksheet.table.Table(displayName=params['table_name'], ref=table_ref)
    out_ws.add_table(table)


    default_sheet_name = 'Sheet'
    if default_sheet_name in out_wb:
        del out_wb[default_sheet_name]

    log.debug(f"saving file { params['out_file_name'] }")
    out_wb.save(params['out_file_name'])


def process_title_row(row, delete_columns):
    """ process an xlrd title row, returning a map of column names to column indexes (origin zero) """
    in_column_map = {}
    out_column_map = {}
    out_col = 0
    for col in range(len(row)):
        value = row[col]
        if value is not None and value != '':

            if value not in delete_columns:
                in_column_map[value] = col
                out_col += 1
                out_column_map[value] = out_col

    return in_column_map, out_column_map



def read_air_travel_roster(session, config, firsttime):

    params0 = {
            'query_id': '481261',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': config.VC_DR_ID,
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = {
            'nd': 'clearreports_auth',
            'init': 'xls',
            'query_id': '481261',
            'prompt0': config.VC_DR_ID,
            'reference': 'disaster',
            }


    return read_common(session, config, params0, params1)



def read_arrival_roster(session, config, firsttime):

    params0 = {
            'query_id': '1537756',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': config.VC_DR_ID,
            'prompt2': 'Arrival Date',
            'prompt3': 'No',
            'prompt4': 'Yes',
            'prompt5': '-1__,__',
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = {
            'nd': 'clearreports_auth',
            'init': 'xls',
            'query_id': '1537756',
            'prompt0': config.VC_DR_ID,
            'prompt1': 'Arrival Date',
            'prompt2': 'No',
            'prompt3': 'Yes',
            'prompt4': "['-1']",
            }


    return read_common(session, config, params0, params1)

def read_open_requests(session, config, firsttime):
    """ read open staff requests """

    params0 = {
            'query_id': '1555803',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': config.VC_DR_ID,
            'output_format': 'xls',
            'run': 'Run',
            }


    params1 = {
            'nd': 'clearreports_auth',
            'init': 'xls',
            'query_id': params0['query_id'],
            'prompt0': params0['prompt1'],
            }

    return read_common(session, config, params0, params1)


def read_shift_tool(session, config, firsttime):
    """ read the dro shift tool query """

    yesterday = TODAY - datetime.timedelta(1)
    nextweek = TODAY + datetime.timedelta(7)

    dr_id = config.VC_DR_ID

    params0 = {
            'query_id': '1737309',
            'nd': 'clearreports_launch_admin',
            'reference': 'disaster',
            'prompt1': f"{ dr_id }__,__",
            'prompt2': yesterday.strftime('%Y-%m-%d'),
            'prompt3': nextweek.strftime('%Y-%m-%d'),
            'prompt4': 'Registered__,__',
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = {
            'nd': 'clearreports_auth',
            'init': 'xls',
            'query_id': params0['query_id'],
            'prompt0': f"['{ dr_id }']",     # dr534-21
            'prompt1': yesterday.strftime('Date(%Y,%m,%d)'),
            'prompt2': nextweek.strftime('Date(%Y,%m,%d)'),
            'prompt3': "['Registered']",
            }

    return read_common(session, config, params0, params1)


def read_staff_roster(session, config, firsttime):
    """ read the staff roster """

    yesterday = TODAY - datetime.timedelta(1)
    nextweek = TODAY + datetime.timedelta(7)

    dr_id = config.VC_DR_ID

    params0 = {
            'query_id': '1537757',
            'reference': 'disaster',
            'nd': 'clearreports_launch_admin',
            'prompt1': f"{ dr_id }",
            'prompt2': 'All',
            'prompt3': 'All',
            'prompt4': 'Name',
            'prompt5': '',
            'prompt6': '',
            'prompt7': 'Yes',
            'prompt8': '-1__,__',
            'prompt9': 'One line per person',
            'output_format': 'xls',
            'run': 'Run',
            }

    params1 = {
            'nd': 'clearreports_auth',
            'init': 'xls',
            'query_id': params0['query_id'],
            'prompt0': params0['prompt1'],
            'prompt1': params0['prompt2'],
            'prompt2': params0['prompt3'],
            'prompt3': params0['prompt4'],
            'prompt4': params0['prompt5'],
            'prompt5': params0['prompt6'],
            'prompt6': params0['prompt7'],
            'prompt7': "['-1']",
            'prompt8': params0['prompt9'],
            }

    return read_common(session, config, params0, params1)

def read_common(session, config, params0, params1):

    url = "https://volunteerconnection.redcross.org"

    headers = {
            #'accept': 'application/json, text/javascript, */*; q=0.01',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'content-type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'DNT': '1',
            'Host': 'volunteerconnection.redcross.org',
            'Origin': 'https://volunteerconnection.redcross.org',
            'Referer': 'https://volunteerconnection.redcross.org/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'sec-gpc': '1',
            'User-Agent': '.Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.75 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            }

    #log.debug(f"params0 { params0 } params1 { params1 }")
    response = session.post(url, data=params0, headers=headers, timeout=config.WEB_TIMEOUT)
    response.raise_for_status()


    response = session.post(url, data=params1, headers=headers, timeout=config.WEB_TIMEOUT)
    response.raise_for_status()

    #log.debug(f"response.content { response.content }")
    anchor = response.html.find('a', first=True)
    href = anchor.attrs['href']

    url2 = url + href

    log.debug(f"url2 { url2 }")

    response = session.get(url2, timeout=config.WEB_TIMEOUT)
    response.raise_for_status()

    log.debug(f"retrieved document.  size is { len(response.text) }, type is '{ response.headers.get('content-type') }'")
    #log.debug(f"document: { response.content }")

    return response.content

def init_config():
    class AttrDict(dict):
        def __init__(self, *args, **kwargs):
            super(AttrDict, self).__init__(*args, **kwargs)
            self.__dict__ = self


    config_dotenv = dotenv.dotenv_values(verbose=True)

    config = AttrDict()
    for item in dir(config_static):
        if not item.startswith('__'):
            config[item] = getattr(config_static, item)

    #log.debug(f"config after copy: { config.keys() }")

    for key, val in config_dotenv.items():
        config[key] = val

    return config

def parse_args():
    parser = argparse.ArgumentParser(
            description="generate and send daily staffing reports",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("--post", help="post to real recipients", action="store_true")
    parser.add_argument("--keep-files", help="don't delete output spreadsheets", action="store_true")

    args = parser.parse_args()
    return args



if __name__ == "__main__":
    main()


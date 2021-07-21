#!/usr/bin/env python

import os
import re
import sys
import logging
import argparse
import datetime

import init_logging

import openpyxl
import openpyxl.utils
import openpyxl.styles
import openpyxl.worksheet.table as table
import openpyxl.styles.colors
import dotenv
import xlrd
import O365

import vc_session
import daily_staffing_reports
import config as config_static
import gen_templates


log = logging.getLogger(__name__)


DATESTAMP = datetime.datetime.now().strftime("%Y-%m-%d")

class AttrDict(dict):
    def __init__(self, *args, **kwargs):
        super(AttrDict, self).__init__(*args, **kwargs)
        self.__dict__ = self

def main():
    args = parse_args()
    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)
    log.debug("running...")

    config = init_config()

    # initialize office 365 graph api
    credentials = (config.CLIENT_ID, config.CLIENT_SECRET)

    scopes = [
            'https://graph.microsoft.com/Files.ReadWrite.All',
            'https://graph.microsoft.com/Mail.Send',
            'https://graph.microsoft.com/Mail.Send.Shared',
            'https://graph.microsoft.com/offline_access',
            'https://graph.microsoft.com/User.Read',
            'https://graph.microsoft.com/User.ReadBasic.All',
            'https://graph.microsoft.com/Sites.ReadWrite.All',
            'https://graph.microsoft.com/offline_access',
            'basic',
            ]

    token_backend = O365.FileSystemTokenBackend(token_path='.', token_filename="my_token.txt")
    account = O365.Account(credentials, scopes=scopes, token_backend=token_backend)

    if not account.is_authenticated:
        account.authenticate()
        if not account.is_authenticated:
            log.fatal(f"Cannot authenticate account")
            sys.exit(1)


    # initialize volunteer connection api

    roster_file = config.ROSTER_FILE
    if args.cached_input:
        if not os.path.exists(roster_file):
            log.fatal(f"Roster file { roster_file } not found")
            sys.exit(1)
        with open(roster_file, "rb") as file:
            data = file.read()
    else:
        session = vc_session.get_session(config)
        data = daily_staffing_reports.read_staff_roster(session, config, args)
        if args.save_input:
            with open(roster_file, "wb") as file:
                file.write(data)


    roster_wb = xlrd.open_workbook(file_contents=data)

    output_wb = make_workbook(roster_wb, config)

    if args.save_output:
        output_file = config.OUTPUT_FILE
        if os.path.exists(output_file):
            os.remove(output_file)

        output_wb.save(output_file)


    templates = gen_templates.init()
    date = datetime.datetime.now().strftime("%Y-%m-%d %H%M")

    if args.send or args.test_send:
        send_mail(account, date, output_wb[config.OUTPUT_SHEET_REPORTING], templates.get_template("mail_supervisor.html"), f"{ config.DR_NAME } Supervisor", config, args)
        send_mail(account, date, output_wb[config.OUTPUT_SHEET_NONSVS],    templates.get_template("mail_nonsvs.html"), config.DR_NAME, config, args)


def send_mail(account, date, ws, template, label, config, args):

    title_row = ws[1]
    title_dict = {}

    col_index = 0
    for col in title_row:
        value = title_row[col_index].value.replace(' ', '_')

        #log.debug(f"title_row: { col_index } = '{ value }'")

        title_dict[col_index] = value
        col_index += 1

    for row_index in range(1, ws.max_row):
        row = ws[row_index + 1]

        col_index = 0
        context = {
                'Date': date,
                'mail_owner': config.MAIL_OWNER,
                'dr_name': config.DR_NAME,
                }
        for col in row:
            value = row[col_index].value
            col_name = title_dict[col_index]
            #log.debug(f"row { row_index }: { col_name } = '{ value }'")

            context[col_name] = value;
            col_index += 1

        expand = template.render(context)
        log.debug(f"message: { expand }")

        name = context['Name']

        m = account.new_message()
        if args.test_send:
            m.bcc.add("generic@askneil.com")
        if args.send:
            m.to.add(context['Email'])
        m.subject = f"{label} VC status - { date } - { name }"
        m.body = expand
        m.send()

        # debugging only
        #if row_index > 3:
        #    break



def make_workbook(roster_wb, config):
    """ common point to generate output workbook from input workbook.

        All file operations are 'above' this function (i.e. done by the callers
        of this function)
    """

    roster_ws = roster_wb.sheet_by_index(0)

    current_row = config.ROSTER_TITLE_ROW

    title_names, title_cols = parse_title_row(roster_ws, current_row)

    sup_dict, no_sups, name_dict = process_roster(roster_ws, current_row, title_names, title_cols)

    output_wb = openpyxl.Workbook()


    # add the reporting info
    generate_sups(output_wb, sup_dict, name_dict, config.OUTPUT_SHEET_REPORTING)

    # add the no_sups folks
    ws = output_wb.create_sheet(title=config.OUTPUT_SHEET_NOSUPS)
    generate_no_sups(ws, no_sups, title_cols)

    # add the non_supervisor folks
    generate_non_svs(output_wb, sup_dict, name_dict, config.OUTPUT_SHEET_NONSVS)

    generate_3days(output_wb, sup_dict, name_dict, config.OUTPUT_SHEET_3DAYS)

    default_sheet_name = 'Sheet'
    if default_sheet_name in output_wb:
        del output_wb[default_sheet_name]

    return output_wb




def parse_title_row(ws, title_row):
    """ grab all the column headers from the title row

        turn them into two dicts: column name > column number and column number > column name
    """
    title_name_dict = {}
    title_col_dict = {}

    row = ws.row_values(title_row -1)

    index = 0
    for value in row:
        #log.debug(f"title { index } = '{ value }'")
        title_name_dict[value] = index
        title_col_dict[index] = value
        index += 1

    return title_name_dict, title_col_dict


def process_roster(ws, start_row, title_names, title_cols):
    """ process all the rows in the worksheet

        ws - the worksheet
        start_row - starting row of data in the ws (origin zero)
        title_names - dict mapping names to column numbers (origin zero)
        title_cols - dict mapping column numbers to column names (origin zero)
    """

    no_sups = []
    sup_dict = {}   # keyed by sup name; contains an array of reporting people
    name_dict = {}  # keyed by name; contains whole list

    nrows = ws.nrows
    row_num = start_row -1 # convert to origin zero
    while row_num < nrows -1:
        row_num += 1

        row = ws.row_values(row_num)
        row_dict = row_to_dict(row, title_cols)

        #log.debug(f"row { row_num } values { row_dict }")

        name = row_dict['Name']
        gap = format_gap(row_dict['GAP(s)'])
        sup = row_dict['Current/Last Supervisor']
        released = row_dict['Released']

        name_dict[name] = row_dict

        # handle folks no longer on the job
        if released != '':
            #log.debug(f"row { row_num } person { name } has been released; skipping...")
            continue


        # handle empty supervisor field
        if sup == "":
            if gap != 'OM//DIR' and gap != 'OM//RCCO':
                no_sups.append(row)
                #log.debug(f"row { row_num } person { name } has no supervisor")
            continue

        # handle first entry in list
        if sup not in sup_dict:
            sup_dict[sup] = []

        sup_dict[sup].append(row_dict)

        # DEBUG ONLY
        #if row_num >= 180:
        #    break

    return sup_dict, no_sups, name_dict


ordinal_1900_01_01 = datetime.datetime(1900, 1, 1).toordinal()
def excel_date_to_string(excel_date):
    """ convert floating excel representation of date to yyyy-mm-dd string """

    if excel_date == "":
        return ""

    dt = excel_to_dt(excel_date)

    return dt.strftime('%Y-%m-%d')

def excel_to_dt(excel_date):

    dt = datetime.datetime.fromordinal(ordinal_1900_01_01 + int(excel_date) -2)

    return dt


gap_pattern = re.compile('.*,')
def format_gap(gap):
    gap = re.sub(gap_pattern,'', gap)
    if gap == '':
        gap = "No Assignment"

    return gap



def format_reports(name, sup_row, name_dict):
    """ format the list of direct reports """

    output = [ f"{'Name':25s} { 'GAP':15s} { 'Checked in':12s} { 'Last Day':12s} {'Cell Phone':14s} { 'Email':30s}" ]

    for row in sup_row:

        #plog.debug(f"name { row['Name'] } released: { row['Released'] }")
        # only add folks still on the DR
        gap = format_gap(row['GAP(s)'])
        checkin = excel_date_to_string(row['Checked in'])
        lastday = excel_date_to_string(row['Expect release'])
        output.append(f"{ row['Name']:25s} { gap:15s} { checkin:12s} { lastday:12s} { row['Cell phone']:14s} { row['Email']:30s}")

    return "\r\n".join(output)


first_regex = re.compile('.*, ')
def get_first(s):
    return re.sub(first_regex, '', s)

last_regex = re.compile(',.*')
def get_last(s):
    return re.sub(last_regex, '', s)

def get_gap(s):
    if s == '':
        s = "No Assignment"

    return s


def get_unknown(s):
    if s == '':
        s = "Unknown"

    return s

def get_supervisor(s):
    if s == '':
        s = "NO SUPERVISOR"

    #log.debug(f"get_supervisor: '{ s }'")
    return s



def generate_sups(wb, sups, name_dict, sheet_name):
    """ generate a sheet for those with direct reports """

    output_cols = [
            { 'name': 'Name', 'width': 20, 'field': lambda x: name_dict[x]['Name'], },
            { 'name': 'First', 'width': 20, 'field': lambda x: get_first(name_dict[x]['Name']), },
            { 'name': 'Last', 'width': 20, 'field': lambda x: get_last(name_dict[x]['Name']), },
            { 'name': 'Email', 'width': 30, 'field': lambda x: name_dict[x]['Email'] },
            { 'name': 'Supervisor', 'width': 30, 'field': lambda x: get_supervisor(name_dict[x]['Current/Last Supervisor']) },
            { 'name': 'Last Day', 'width': 30, 'field': lambda x: excel_date_to_string(name_dict[x]['Expect release']) },
            { 'name': 'GAP', 'width': 30, 'field': lambda x: format_gap(name_dict[x]['GAP(s)']) },
            { 'name': 'Location type', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Location type']) },
            { 'name': 'District', 'width': 30, 'field': lambda x: name_dict[x]['District'] },
            { 'name': 'Work Location', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Reporting/Work Location']) },
            { 'name': 'Current lodging', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Current lodging']) },
            { 'name': 'Reports', 'width': 50, 'field': lambda x: format_reports(x, sups[x], name_dict), },
            ]

    generate_sheet(wb, sups, output_cols, sheet_name)

def generate_non_svs(wb, sups, name_dict, sheet_name):
    """ generate a sheet for those without direct reports """


    # get all the folks on name_dict who aren't on sups
    individuals = {}
    for name, val in name_dict.items():

        # skip people on the supervisor lsit
        if name in sups:
            continue

        # skip people not checked in
        if val['Checked in'] == "":
            continue

        # skip people outprocessed
        if val['Released'] != "":
            continue

        individuals[name] = val

    output_cols = [
            { 'name': 'Name', 'width': 20, 'field': lambda x: name_dict[x]['Name'], },
            { 'name': 'First', 'width': 20, 'field': lambda x: get_first(name_dict[x]['Name']), },
            { 'name': 'Last', 'width': 20, 'field': lambda x: get_last(name_dict[x]['Name']), },
            { 'name': 'Email', 'width': 30, 'field': lambda x: name_dict[x]['Email'] },
            { 'name': 'Supervisor', 'width': 30, 'field': lambda x: get_supervisor(name_dict[x]['Current/Last Supervisor']) },
            { 'name': 'Last Day', 'width': 30, 'field': lambda x: excel_date_to_string(name_dict[x]['Expect release']) },
            { 'name': 'GAP', 'width': 30, 'field': lambda x: format_gap(name_dict[x]['GAP(s)']) },
            { 'name': 'Work Location', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Reporting/Work Location']) },
            { 'name': 'Location type', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Location type']) },
            { 'name': 'Current lodging', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Current lodging']) },
            ]

    generate_sheet(wb, individuals, output_cols, sheet_name)


def generate_3days(wb, sups, name_dict, sheet_name):
    """ add everyone with 3 or fewer days left on the job """


    short = {}
    three_days = datetime.datetime.now() + datetime.timedelta(days=3)    # 3 days in the future
    for name, val in name_dict.items():

        # skip people not checked in
        if val['Checked in'] == "":
            continue

        # skip people outprocessed
        if val['Released'] != "":
            continue

        out_date = excel_to_dt(val['Expect release'])
        if out_date > three_days:
            continue

        short[name] = val

    output_cols = [
            { 'name': 'Name', 'width': 20, 'field': lambda x: name_dict[x]['Name'], },
            { 'name': 'First', 'width': 20, 'field': lambda x: get_first(name_dict[x]['Name']), },
            { 'name': 'Last', 'width': 20, 'field': lambda x: get_last(name_dict[x]['Name']), },
            { 'name': 'Email', 'width': 30, 'field': lambda x: name_dict[x]['Email'] },
            { 'name': 'Supervisor', 'width': 30, 'field': lambda x: get_supervisor(name_dict[x]['Current/Last Supervisor']) },
            { 'name': 'Last Day', 'width': 30, 'field': lambda x: excel_date_to_string(name_dict[x]['Expect release']) },
            { 'name': 'GAP', 'width': 30, 'field': lambda x: format_gap(name_dict[x]['GAP(s)']) },
            { 'name': 'Work Location', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Reporting/Work Location']) },
            { 'name': 'Location type', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Location type']) },
            { 'name': 'Current lodging', 'width': 30, 'field': lambda x: get_unknown(name_dict[x]['Current lodging']) },
            ]

    generate_sheet(wb, short, output_cols, sheet_name)




def generate_sheet(wb, people, output_cols, sheet_name):
    """ generate a row for each supervisor, with their direct reports and contact info """

    ws = wb.create_sheet(title=sheet_name)

    # generate the title row
    for i, col_def in enumerate(output_cols):
        ws.cell(column=i+1, row=1, value=col_def['name'])
        width = col_def['width']
        #log.debug(f"column { col_def['name'] } has width { width }")
        ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = col_def['width']


    output_row = 1
    for name in sorted(people.keys()):
        output_row += 1
        sup_row = people[name]

        #log.debug(f"processing sup { name } len { len(sup_row) }")

        for i, col_def in enumerate(output_cols):
            value = col_def['field'](name)
            ws.cell(column=i+1, row=output_row, value=value)

    ref = f"A1:{openpyxl.utils.get_column_letter(len(output_cols))}{ output_row }"
    log.debug(f"sheet { sheet_name } table range ref is '{ ref }'")

    tab = table.Table(displayName=sheet_name, ref=ref)
    ws.add_table(tab)


def generate_no_sups(ws, no_sups, col_dict):
    """ add the no_sups info to the worksheet

        make sure to convert from origin 0 input to origin 1 output

    """

    date_column_names = { 'Assigned':1, 'Checked in': 1, 'Expect release': 1, }
    date_column_cols = {}

    #log.debug(f"col_dict { col_dict }")

    # convert dicts back to array
    column_array = []
    column_index = 0
    for key in sorted(col_dict.keys()):
        column_index += 1
        name = col_dict[key]
        if name in date_column_names:
            date_column_cols[key] = 1
            column_letter = openpyxl.utils.cell.get_column_letter(column_index)
            ws.column_dimensions[column_letter].width = 14
        column_array.append(name)

    # fill in the title row
    for col, val in enumerate(column_array):
        ws.cell(column=col+1, row=1, value=val)
        #log.debug(f"setting title col { col + 1 } to { val }")

    row_num = 1
    for row in no_sups:
        row_num += 1
        col_num = -1
        for col in row:
            col_num += 1
            out_cell = ws.cell(column=col_num + 1, row=row_num, value=col)

            # make sure date columns are formatted properly
            if col_num in date_column_cols:
                out_cell.number_format = 'yyyy-mm-dd'




def row_to_dict(row, title_cols):
    """ convert a row of data into a dict mapping column name to value """

    result = {}
    for index, value in enumerate(row):
        result[title_cols[index]] = value

    return result


def init_config():
    class AttrDict(dict):
        def __init__(self, *args, **kwargs):
            super(AttrDict, self).__init__(*args, **kwargs)
            self.__dict__ = self


    config_dotenv = dotenv.dotenv_values(verbose=True)

    config = AttrDict()
    for item in dir(config_static):
        if not item.startswith('__'):
            config[item] = getattr(config_static, item)

    #log.debug(f"config after copy: { config.keys() }")

    for key, val in config_dotenv.items():
        config[key] = val

    return config


def parse_args():
    parser = argparse.ArgumentParser(
            description="Organize staff roster for mailing reports",
            allow_abbrev=False)
    parser.add_argument("--debug", help="turn on debugging output", action="store_true")
    parser.add_argument("--send", help="send emails to folks on the DR", action="store_true")
    parser.add_argument("--test-send", help="send emails to a test recipient", action="store_true")
    parser.add_argument("--save-input", help="save a copy of the raw VC spreadsheets", action="store_true")
    parser.add_argument("--cached-input", help="use the saved copies of the raw VC spreadsheets", action="store_true")
    parser.add_argument("--save-output", help="don't delete output spreadsheets", action="store_true")

    args = parser.parse_args()
    return args


if __name__ == "__main__":
    main()

